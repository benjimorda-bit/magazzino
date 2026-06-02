"""
Gestionale Magazzino Tappeti
Applicazione desktop per la gestione di un magazzino tappeti.
Sviluppato con customtkinter, pandas, openpyxl e tksheet.
"""

import os
import re
import sys
from datetime import datetime, timedelta
from tkinter import messagebox, filedialog
import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tksheet import Sheet

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


def _get_percorso_database() -> str:
    """Restituisce il percorso assoluto del file database,
    corretto sia in esecuzione normale che come .exe PyInstaller."""
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "magazzino_tappeti.xlsx")


FILE_DATABASE = _get_percorso_database()

COLONNE_DATABASE = [
    "Stato", "Nr", "Nome", "Provenienza", "Misura", "MQ", "UM",
    "Epoca", "Qualita", "Disegno", "Colore", "Fornitore",
    "Costo", "Listino", "Prezzo Vendita Effettivo", "Data Vendita", "Note"
]
COLONNE_NUMERICHE = ["MQ", "Costo", "Listino", "Prezzo Vendita Effettivo"]
OPZIONI_STATO = ["Disponibile", "Venduto"]
OPZIONI_EPOCA = ["Nuovo", "Vecchio", "Antico"]
OPZIONI_QUALITA = ["Medio", "Fine", "Commerciale", "Extra"]
OPZIONI_DISEGNO = ["Etnico", "Classico", "Decorativo", "Moderno"]
OPZIONI_PROVENIENZA = [
    "Iran", "Turchia", "Caucaso", "Russia", "Cina",
    "Nepal", "Afghanistan", "India", "Pakistan", "Marocco"
]
OPZIONI_UM = [
    "corsia", "corsietta", "pedana", "zarunim", "sajade",
    "parde", "250x2", "3x2", "3x250", "350x250", "4x3",
    "grande", "kelley", "tondo"
]

LARGHEZZE_COLONNE = {
    "Stato": 90, "Nr": 65, "Nome": 155, "Provenienza": 100,
    "Misura": 90, "MQ": 60, "UM": 55, "Epoca": 70, "Qualita": 90,
    "Disegno": 85, "Colore": 80, "Fornitore": 110,
    "Costo": 90, "Listino": 90, "Prezzo Vendita Effettivo": 145,
    "Data Vendita": 125, "Note": 150
}


def formatta_prezzo(valore: float) -> str:
    """Formatta un numero come prezzo in stile italiano.
    3500 → '3.500'   |   6789.82 → '6.789,82'   |   0 → ''
    """
    if not valore:
        return ""
    arrotondato = round(float(valore), 2)
    if arrotondato == int(arrotondato):
        return f"{int(arrotondato):,}".replace(",", ".")
    s = f"{arrotondato:,.2f}"          # es. "6,789.82"
    parti = s.split(".")
    return parti[0].replace(",", ".") + "," + parti[1]


def _parse_numero(val_str: str) -> float:
    """Converte una stringa di prezzo in float.
    Gestisce formato italiano (3.500,50) ed europeo/americano (3500.50).
    """
    s = re.sub(r"[^\d.,]", "", str(val_str).strip())
    if not s:
        return 0.0
    if "," in s and "." in s:
        # Formato italiano: punto = migliaia, virgola = decimale
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    elif "." in s:
        parti = s.split(".")
        # Se tutte le parti dopo il primo punto hanno 3 cifre → separatore migliaia
        if len(parti) >= 2 and all(len(p) == 3 for p in parti[1:]):
            s = s.replace(".", "")
        # altrimenti è un decimale normale (es. "3500.50")
    try:
        return float(s)
    except ValueError:
        return 0.0


def formatta_excel_per_stampa(percorso: str, totali: dict = None):
    """Applica al file Excel uno stile pronto per la stampa:
    bordi su tutte le celle, header in grassetto con sfondo grigio chiaro,
    larghezza colonne auto-regolata, griglia di stampa attiva, e
    box totali in fondo (se fornito)."""
    try:
        wb = load_workbook(percorso)
        ws = wb.active

        # Bordi sottili su tutte le celle dati
        lato = Side(border_style="thin", color="000000")
        bordo = Border(left=lato, right=lato, top=lato, bottom=lato)
        ultima_riga_dati = ws.max_row
        ultima_colonna_dati = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=ultima_riga_dati,
                                 min_col=1, max_col=ultima_colonna_dati):
            for cell in row:
                cell.border = bordo
                cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Header in grassetto con sfondo grigio chiaro
        grassetto_header = Font(bold=True, size=11)
        sfondo_header = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        for cell in ws[1]:
            cell.font = grassetto_header
            cell.fill = sfondo_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-larghezza colonne (basata sul contenuto, con minimo e massimo)
        for col_cells in ws.columns:
            lunghezze = []
            for cell in col_cells:
                if cell.value is not None:
                    lunghezze.append(len(str(cell.value)))
            if lunghezze:
                massimo = max(lunghezze)
                larghezza = min(max(massimo + 2, 8), 40)
                ws.column_dimensions[col_cells[0].column_letter].width = larghezza

        # Totali come piè di pagina (visibile solo in stampa, su una riga,
        # in basso, staccato dai dati). Appare in fondo a OGNI pagina stampata.
        if totali:
            totali_str = (
                f"Pezzi totali: {totali.get('pezzi', 0)}     "
                f"MQ totali: {totali.get('mq', 0):.2f}     "
                f"Costo totale: {formatta_prezzo(totali.get('costo', 0)) or '0'} €     "
                f"Listino totale: {formatta_prezzo(totali.get('listino', 0)) or '0'} €     "
                f"Incasso totale: {formatta_prezzo(totali.get('incasso', 0)) or '0'} €"
            )
            ws.oddFooter.center.text = totali_str
            ws.oddFooter.center.size = 11
            ws.oddFooter.center.font = "Arial,Bold"

        # Impostazioni di stampa: griglia visibile, orientamento orizzontale
        ws.print_options.gridLines = True
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:1"  # ripete l'header su ogni pagina stampata

        wb.save(percorso)
    except Exception:
        # Non bloccare l'export se la formattazione fallisce
        pass


def _chiave_naturale_nr(valore) -> list:
    """Chiave di ordinamento 'naturale' per il codice Nr.
    Funziona sia con codici puramente numerici (es. 216, 224, 226)
    che con codici alfanumerici (es. T001, T010, T100)."""
    s = str(valore).strip()
    return [int(t) if t.isdigit() else t.lower()
            for t in re.split(r"(\d+)", s)]


def ordina_df_per_nr(df: pd.DataFrame) -> pd.DataFrame:
    """Ordina un dataframe per Nr in ordine crescente naturale."""
    if df.empty:
        return df
    indici_ordinati = sorted(
        df.index,
        key=lambda i: _chiave_naturale_nr(df.loc[i, "Nr"])
    )
    return df.loc[indici_ordinati].reset_index(drop=True)


class AutocompleteEntry(ctk.CTkFrame):
    """Entry con tendina di autocompletamento in tempo reale."""

    def __init__(self, parent, width: int = 380,
                 placeholder: str = "Scrivi Nr o nome...", on_select=None):
        super().__init__(parent, fg_color="transparent")
        self._values = []
        self._on_select = on_select

        self._entry = ctk.CTkEntry(
            self, width=width, placeholder_text=placeholder
        )
        self._entry.pack(fill="x")
        self._entry.bind("<KeyRelease>", self._filtra)
        self._entry.bind("<Escape>", lambda e: self._nascondi())

        self._frame_dropdown = ctk.CTkFrame(self, corner_radius=6)
        self._visibile = False

    def _filtra(self, event=None):
        termine = self._entry.get().strip().lower()
        if not termine:
            self._nascondi()
            return
        corrispondenze = [v for v in self._values if termine in v.lower()]
        if corrispondenze:
            self._mostra(corrispondenze)
        else:
            self._nascondi()

    def _mostra(self, items):
        for w in self._frame_dropdown.winfo_children():
            w.destroy()
        for item in items[:8]:
            ctk.CTkButton(
                self._frame_dropdown,
                text=item, anchor="w",
                fg_color="transparent",
                text_color=("gray10", "gray90"),
                hover_color=("gray85", "gray30"),
                height=30,
                command=lambda i=item: self._seleziona(i)
            ).pack(fill="x", padx=4, pady=1)
        if not self._visibile:
            self._frame_dropdown.pack(fill="x", pady=(2, 0))
            self._visibile = True

    def _nascondi(self):
        if self._visibile:
            self._frame_dropdown.pack_forget()
            self._visibile = False

    def _seleziona(self, valore: str):
        self._entry.delete(0, "end")
        self._entry.insert(0, valore)
        self._nascondi()
        if self._on_select:
            self._on_select(valore)

    def set_values(self, values: list):
        self._values = values

    def get(self) -> str:
        return self._entry.get().strip()

    def set(self, value: str):
        self._entry.delete(0, "end")
        self._entry.insert(0, value or "")
        self._nascondi()


class GestoreDatabase:
    """Classe per la gestione del database Excel."""

    def __init__(self, percorso_file: str):
        self.percorso_file = percorso_file
        self.df = self._carica_database()

    def _carica_database(self) -> pd.DataFrame:
        """Carica il database da file Excel o ne crea uno nuovo."""
        if os.path.exists(self.percorso_file):
            try:
                df = pd.read_excel(self.percorso_file, engine="openpyxl")
                for col in COLONNE_DATABASE:
                    if col not in df.columns:
                        df[col] = ""
                df = df[COLONNE_DATABASE]
            except Exception as e:
                messagebox.showerror("Errore", f"Errore nel caricamento del database:\n{e}")
                df = pd.DataFrame(columns=COLONNE_DATABASE)
        else:
            df = pd.DataFrame(columns=COLONNE_DATABASE)

        df["Nr"] = df["Nr"].astype(str)
        df = self._forza_colonne_numeriche(df)
        return df

    def _forza_colonne_numeriche(self, df: pd.DataFrame) -> pd.DataFrame:
        """Converte le colonne monetarie e MQ in valori numerici."""
        for col in COLONNE_NUMERICHE:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df

    def salva_database(self) -> bool:
        """Salva il dataframe nel file Excel."""
        try:
            self.df.to_excel(self.percorso_file, index=False, engine="openpyxl")
            return True
        except Exception as e:
            messagebox.showerror("Errore", f"Errore nel salvataggio:\n{e}")
            return False

    def aggiungi_tappeto(self, dati: dict) -> bool:
        """Aggiunge un nuovo tappeto al database."""
        if str(dati["Nr"]) in self.df["Nr"].values:
            messagebox.showwarning("Attenzione", "Il codice inserito esiste gia nel database.")
            return False
        dati["Nr"] = str(dati["Nr"])
        nuova_riga = pd.DataFrame([dati])
        nuova_riga = self._forza_colonne_numeriche(nuova_riga)
        self.df = pd.concat([self.df, nuova_riga], ignore_index=True)
        return self.salva_database()

    def elimina_tappeto(self, codice: str) -> bool:
        """Elimina un tappeto dal database tramite codice."""
        self.df = self.df[self.df["Nr"].astype(str) != str(codice)]
        return self.salva_database()

    def registra_vendita(self, codice: str, prezzo_effettivo: float) -> bool:
        """Registra la vendita di un tappeto."""
        idx = self.df[self.df["Nr"].astype(str) == str(codice)].index
        if len(idx) == 0:
            return False
        self.df.loc[idx, "Stato"] = "Venduto"
        self.df.loc[idx, "Prezzo Vendita Effettivo"] = prezzo_effettivo
        self.df.loc[idx, "Data Vendita"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        return self.salva_database()

    def azzera_storico_vendite(self) -> bool:
        """Elimina tutti i tappeti venduti dal database."""
        self.df = self.df[self.df["Stato"] != "Venduto"]
        return self.salva_database()

    def filtra_dati(self, filtri: dict) -> pd.DataFrame:
        """Filtra i dati in base ai criteri specificati."""
        df_filtrato = self.df.copy()

        if filtri.get("nome_codice"):
            termine = filtri["nome_codice"].lower()
            mask = (
                df_filtrato["Nr"].astype(str).str.lower().str.contains(termine, na=False) |
                df_filtrato["Nome"].astype(str).str.lower().str.contains(termine, na=False)
            )
            df_filtrato = df_filtrato[mask]

        if filtri.get("provenienza"):
            termine = filtri["provenienza"].lower()
            df_filtrato = df_filtrato[
                df_filtrato["Provenienza"].astype(str).str.lower().str.contains(termine, na=False)
            ]

        if filtri.get("fornitore"):
            termine = filtri["fornitore"].lower()
            df_filtrato = df_filtrato[
                df_filtrato["Fornitore"].astype(str).str.lower().str.contains(termine, na=False)
            ]

        if filtri.get("um"):
            termine = filtri["um"].lower()
            df_filtrato = df_filtrato[
                df_filtrato["UM"].astype(str).str.lower().str.contains(termine, na=False)
            ]

        if filtri.get("stato") and filtri["stato"] != "Tutti":
            df_filtrato = df_filtrato[df_filtrato["Stato"] == filtri["stato"]]

        # Ordina sempre per Nr crescente
        return ordina_df_per_nr(df_filtrato)

    def ottieni_disponibili(self) -> pd.DataFrame:
        """Restituisce solo i tappeti disponibili."""
        return self.df[self.df["Stato"] == "Disponibile"]

    def ottieni_venduti_periodo(self, giorni: int = None) -> pd.DataFrame:
        """Restituisce i tappeti venduti in un determinato periodo."""
        venduti = self.df[self.df["Stato"] == "Venduto"].copy()

        if giorni is None or venduti.empty:
            return venduti

        venduti["Data Vendita"] = pd.to_datetime(venduti["Data Vendita"], errors="coerce")
        data_limite = datetime.now() - timedelta(days=giorni)
        return venduti[venduti["Data Vendita"] >= data_limite]


def calcola_mq_da_misura(misura: str) -> float:
    """Calcola i metri quadri dalla stringa misura (es. '299*74' o '299x74')."""
    if not misura:
        return 0.0

    pattern = r"(\d+(?:[.,]\d+)?)\s*[xX*]\s*(\d+(?:[.,]\d+)?)"
    match = re.search(pattern, str(misura))

    if match:
        try:
            lunghezza = float(match.group(1).replace(",", "."))
            larghezza = float(match.group(2).replace(",", "."))
            return round((lunghezza * larghezza) / 10000, 2)
        except ValueError:
            return 0.0
    return 0.0


class FrameDashboard(ctk.CTkFrame):
    """Frame per la sezione Dashboard."""

    def __init__(self, master, gestore_db: GestoreDatabase):
        super().__init__(master, fg_color="transparent")
        self.gestore_db = gestore_db
        self._crea_interfaccia()

    def _crea_card(self, parent, titolo: str, col: int) -> ctk.CTkLabel:
        frame = ctk.CTkFrame(parent, corner_radius=10)
        frame.grid(row=0, column=col, padx=8, pady=8, sticky="ew")
        parent.columnconfigure(col, weight=1)
        ctk.CTkLabel(
            frame, text=titolo, font=ctk.CTkFont(size=11), text_color="gray"
        ).pack(pady=(12, 2), padx=15)
        lbl = ctk.CTkLabel(frame, text="—", font=ctk.CTkFont(size=16, weight="bold"))
        lbl.pack(pady=(2, 12), padx=15)
        return lbl

    def _crea_card_stat(self, parent, titolo: str, col: int,
                        colore: str = None) -> ctk.CTkLabel:
        frame = ctk.CTkFrame(parent, corner_radius=8)
        frame.grid(row=0, column=col, padx=6, pady=6, sticky="ew")
        parent.columnconfigure(col, weight=1)
        ctk.CTkLabel(
            frame, text=titolo, font=ctk.CTkFont(size=10), text_color="gray"
        ).pack(pady=(8, 1), padx=12)
        kwargs = {"font": ctk.CTkFont(size=14, weight="bold")}
        if colore:
            kwargs["text_color"] = colore
        lbl = ctk.CTkLabel(frame, text="—", **kwargs)
        lbl.pack(pady=(1, 8), padx=12)
        return lbl

    def _crea_interfaccia(self):
        """Crea tutti gli elementi dell'interfaccia dashboard."""
        ctk.CTkLabel(
            self, text="Dashboard", font=ctk.CTkFont(size=24, weight="bold")
        ).pack(pady=(20, 10))

        frame_resoconto = ctk.CTkFrame(self)
        frame_resoconto.pack(fill="x", padx=20, pady=10)

        ctk.CTkLabel(
            frame_resoconto, text="Resoconto Magazzino Attuale",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(pady=(12, 8))

        frame_cards = ctk.CTkFrame(frame_resoconto, fg_color="transparent")
        frame_cards.pack(pady=(0, 10), padx=20, fill="x")

        self.label_pezzi = self._crea_card(frame_cards, "Pezzi Disponibili", 0)
        self.label_mq_tot = self._crea_card(frame_cards, "MQ Totali Disponibili", 1)
        self.label_costo = self._crea_card(frame_cards, "Costo Totale", 2)
        self.label_listino = self._crea_card(frame_cards, "Listino Totale", 3)

        ctk.CTkButton(
            frame_resoconto,
            text="Azzera Storico Vendite (Fine Anno)",
            fg_color="#dc3545", hover_color="#c82333",
            command=self._azzera_storico
        ).pack(pady=(0, 15))

        ctk.CTkLabel(
            self, text="Statistiche Vendite",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(pady=(20, 10))

        self.tabview = ctk.CTkTabview(self, width=500, height=220)
        self.tabview.pack(padx=20, pady=10, fill="x")

        self.tabs = {
            "Ultimo Mese": 30,
            "Ultimi 6 Mesi": 180,
            "Ultimo Anno": 365,
            "Sempre (Totale)": None
        }

        self.labels_stats = {}
        for nome_tab in self.tabs.keys():
            tab = self.tabview.add(nome_tab)
            frame_tab = ctk.CTkFrame(tab, fg_color="transparent")
            frame_tab.pack(fill="both", expand=True, padx=10, pady=10)
            frame_tab_cards = ctk.CTkFrame(frame_tab, fg_color="transparent")
            frame_tab_cards.pack(fill="x")

            self.labels_stats[nome_tab] = {
                "pezzi":   self._crea_card_stat(frame_tab_cards, "Pezzi Venduti", 0),
                "incasso": self._crea_card_stat(frame_tab_cards, "Incasso Totale", 1),
                "costo":   self._crea_card_stat(frame_tab_cards, "Costo Merci", 2),
                "utile":   self._crea_card_stat(frame_tab_cards, "Utile Netto", 3, "#2d8a4e"),
            }

        self.aggiorna_dati()

    def _azzera_storico(self):
        """Azzera lo storico delle vendite dopo conferma."""
        risposta = messagebox.askyesno(
            "Conferma Azzeramento",
            "Sei sicuro di voler eliminare definitivamente tutti i tappeti venduti?\n\n"
            "Questa operazione non puo essere annullata."
        )
        if risposta:
            if self.gestore_db.azzera_storico_vendite():
                messagebox.showinfo("Completato", "Storico vendite azzerato con successo.")
                self.aggiorna_dati()

    def aggiorna_dati(self):
        """Aggiorna tutti i dati della dashboard."""
        disponibili = self.gestore_db.ottieni_disponibili()
        self.label_pezzi.configure(text=str(len(disponibili)))
        self.label_mq_tot.configure(text=f"{disponibili['MQ'].sum():.2f} m²")
        self.label_costo.configure(
            text=formatta_prezzo(disponibili["Costo"].sum()) or "0"
        )
        self.label_listino.configure(
            text=formatta_prezzo(disponibili["Listino"].sum()) or "0"
        )

        for nome_tab, giorni in self.tabs.items():
            venduti = self.gestore_db.ottieni_venduti_periodo(giorni)
            pezzi = len(venduti)
            incasso = venduti["Prezzo Vendita Effettivo"].sum() if not venduti.empty else 0
            costo = venduti["Costo"].sum() if not venduti.empty else 0
            utile = incasso - costo

            self.labels_stats[nome_tab]["pezzi"].configure(text=str(pezzi))
            self.labels_stats[nome_tab]["incasso"].configure(
                text=formatta_prezzo(incasso) or "0"
            )
            self.labels_stats[nome_tab]["costo"].configure(
                text=formatta_prezzo(costo) or "0"
            )
            utile_colore = "#2d8a4e" if utile >= 0 else "#dc3545"
            self.labels_stats[nome_tab]["utile"].configure(
                text=formatta_prezzo(abs(utile)) or "0",
                text_color=utile_colore
            )


class FrameCercaElenco(ctk.CTkFrame):
    """Frame per la sezione Cerca/Elenco."""

    def __init__(self, master, gestore_db: GestoreDatabase):
        super().__init__(master, fg_color="transparent")
        self.gestore_db = gestore_db
        self.df_visualizzato = pd.DataFrame()
        self._crea_interfaccia()

    def _crea_interfaccia(self):
        """Crea tutti gli elementi dell'interfaccia cerca/elenco."""
        ctk.CTkLabel(
            self, text="Cerca / Elenco Tappeti",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(pady=(20, 10))

        frame_filtri = ctk.CTkFrame(self)
        frame_filtri.pack(fill="x", padx=20, pady=10)

        frame_riga1 = ctk.CTkFrame(frame_filtri, fg_color="transparent")
        frame_riga1.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(frame_riga1, text="Nome/Codice:").pack(side="left", padx=5)
        self.entry_nome_codice = ctk.CTkEntry(
            frame_riga1, width=150, placeholder_text="Cerca..."
        )
        self.entry_nome_codice.pack(side="left", padx=5)
        self.entry_nome_codice.bind("<Return>", lambda e: self._applica_filtri())

        ctk.CTkLabel(frame_riga1, text="Provenienza:").pack(side="left", padx=5)
        self.entry_provenienza = ctk.CTkEntry(frame_riga1, width=120)
        self.entry_provenienza.pack(side="left", padx=5)
        self.entry_provenienza.bind("<Return>", lambda e: self._applica_filtri())

        ctk.CTkLabel(frame_riga1, text="Fornitore:").pack(side="left", padx=5)
        self.entry_fornitore = ctk.CTkEntry(frame_riga1, width=120)
        self.entry_fornitore.pack(side="left", padx=5)
        self.entry_fornitore.bind("<Return>", lambda e: self._applica_filtri())

        frame_riga2 = ctk.CTkFrame(frame_filtri, fg_color="transparent")
        frame_riga2.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(frame_riga2, text="UM:").pack(side="left", padx=5)
        self.entry_um = ctk.CTkEntry(frame_riga2, width=100)
        self.entry_um.pack(side="left", padx=5)
        self.entry_um.bind("<Return>", lambda e: self._applica_filtri())

        ctk.CTkLabel(frame_riga2, text="Stato:").pack(side="left", padx=5)
        self.combo_stato = ctk.CTkComboBox(
            frame_riga2, values=["Tutti"] + OPZIONI_STATO, width=120
        )
        self.combo_stato.set("Tutti")
        self.combo_stato.pack(side="left", padx=5)

        ctk.CTkButton(
            frame_riga2, text="Applica Filtri", command=self._applica_filtri, width=120
        ).pack(side="left", padx=15)

        ctk.CTkButton(
            frame_riga2, text="Azzera", command=self._azzera_filtri, width=80
        ).pack(side="left", padx=5)

        frame_tabella = ctk.CTkFrame(self)
        frame_tabella.pack(fill="both", expand=True, padx=20, pady=10)

        self.sheet = Sheet(
            frame_tabella,
            headers=COLONNE_DATABASE,
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            width=900,
            height=400
        )
        self.sheet.pack(fill="both", expand=True)

        # Abilita la modifica diretta delle celle
        self.sheet.enable_bindings(
            "single_select", "row_select", "column_width_resize", "edit_cell"
        )
        # Intercetta la fine della modifica per aggiornare il database
        try:
            self.sheet.extra_bindings("end_edit_cell", func=self._on_fine_modifica)
        except Exception:
            self.sheet.extra_bindings([("end_edit_cell", self._on_fine_modifica)])

        frame_inferiore = ctk.CTkFrame(self)
        frame_inferiore.pack(fill="x", padx=20, pady=10)

        frame_pulsanti = ctk.CTkFrame(frame_inferiore, fg_color="transparent")
        frame_pulsanti.pack(side="left")

        ctk.CTkButton(
            frame_pulsanti, text="Elimina Selezionato",
            fg_color="#dc3545", hover_color="#c82333",
            command=self._elimina_selezionato
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            frame_pulsanti, text="Esporta per Stampa",
            command=self._esporta_stampa
        ).pack(side="left", padx=5)

        frame_contatori = ctk.CTkFrame(frame_inferiore, fg_color="transparent")
        frame_contatori.pack(side="right")

        self.label_totale_pezzi = ctk.CTkLabel(
            frame_contatori, text="Pezzi: 0", font=ctk.CTkFont(weight="bold")
        )
        self.label_totale_pezzi.pack(side="left", padx=12)

        self.label_totale_mq = ctk.CTkLabel(
            frame_contatori, text="MQ Tot: 0.00",
            font=ctk.CTkFont(weight="bold"), text_color="#1f78b4"
        )
        self.label_totale_mq.pack(side="left", padx=12)

        self.label_totale_costo = ctk.CTkLabel(
            frame_contatori, text="Costo: 0", font=ctk.CTkFont(weight="bold")
        )
        self.label_totale_costo.pack(side="left", padx=12)

        self.label_totale_listino = ctk.CTkLabel(
            frame_contatori, text="Listino: 0", font=ctk.CTkFont(weight="bold")
        )
        self.label_totale_listino.pack(side="left", padx=12)

        self.aggiorna_dati()

    def _on_fine_modifica(self, event):
        """Aggiorna il database dopo la modifica diretta di una cella."""
        try:
            riga_idx = getattr(event, "row", None)
            col_idx = getattr(event, "column", None)
            nuovo_valore = getattr(event, "value", None)
            if riga_idx is None and hasattr(event, "__getitem__"):
                riga_idx = event["row"]
                col_idx = event["column"]
                nuovo_valore = event.get("value", "")
        except Exception:
            return

        if riga_idx is None or col_idx is None:
            return
        if riga_idx < 0 or riga_idx >= len(self.df_visualizzato):
            return

        col_nome = COLONNE_DATABASE[col_idx]
        codice = str(self.df_visualizzato.iloc[riga_idx]["Nr"])

        idx_db = self.gestore_db.df[
            self.gestore_db.df["Nr"].astype(str) == codice
        ].index
        if len(idx_db) == 0:
            return

        # Aggiorna il valore nel database principale
        if col_nome in COLONNE_NUMERICHE:
            valore_parsed = _parse_numero(nuovo_valore)
            self.gestore_db.df.loc[idx_db, col_nome] = valore_parsed
        else:
            self.gestore_db.df.loc[idx_db, col_nome] = str(nuovo_valore)

        # Se viene modificata la Misura, ricalcola i MQ automaticamente
        if col_nome == "Misura":
            nuovi_mq = calcola_mq_da_misura(str(nuovo_valore))
            self.gestore_db.df.loc[idx_db, "MQ"] = nuovi_mq

        self.gestore_db.salva_database()

        # Aggiorna la tabella mantenendo i filtri attivi
        self._applica_filtri()

    def _applica_filtri(self):
        """Applica i filtri e aggiorna la visualizzazione."""
        filtri = {
            "nome_codice": self.entry_nome_codice.get().strip(),
            "provenienza": self.entry_provenienza.get().strip(),
            "fornitore":   self.entry_fornitore.get().strip(),
            "um":          self.entry_um.get().strip(),
            "stato":       self.combo_stato.get()
        }
        self.df_visualizzato = self.gestore_db.filtra_dati(filtri)
        self._aggiorna_tabella()

    def _azzera_filtri(self):
        """Azzera tutti i filtri."""
        self.entry_nome_codice.delete(0, "end")
        self.entry_provenienza.delete(0, "end")
        self.entry_fornitore.delete(0, "end")
        self.entry_um.delete(0, "end")
        self.combo_stato.set("Tutti")
        self.aggiorna_dati()

    def _aggiorna_tabella(self):
        """Aggiorna la visualizzazione della tabella."""
        df_display = self.df_visualizzato.copy()

        # Prezzi: formato italiano senza EUR
        for col in ["Costo", "Listino", "Prezzo Vendita Effettivo"]:
            if col in df_display.columns:
                df_display[col] = df_display[col].apply(
                    lambda x: formatta_prezzo(x) if pd.notna(x) and x != 0 else ""
                )

        if "MQ" in df_display.columns:
            df_display["MQ"] = df_display["MQ"].apply(
                lambda x: f"{x:.2f}" if pd.notna(x) and x != 0 else ""
            )

        dati = df_display.values.tolist()
        self.sheet.set_sheet_data(dati)

        try:
            for i, col in enumerate(COLONNE_DATABASE):
                self.sheet.column_width(column=i, width=LARGHEZZE_COLONNE.get(col, 100))
        except Exception:
            self.sheet.set_all_column_widths(100)

        pezzi = len(self.df_visualizzato)
        mq_tot = self.df_visualizzato["MQ"].sum() if not self.df_visualizzato.empty else 0
        costo_tot = self.df_visualizzato["Costo"].sum() if not self.df_visualizzato.empty else 0
        listino_tot = self.df_visualizzato["Listino"].sum() if not self.df_visualizzato.empty else 0

        self.label_totale_pezzi.configure(text=f"Pezzi: {pezzi}")
        self.label_totale_mq.configure(text=f"MQ Tot: {mq_tot:.2f}")
        self.label_totale_costo.configure(
            text=f"Costo: {formatta_prezzo(costo_tot) or '0'}"
        )
        self.label_totale_listino.configure(
            text=f"Listino: {formatta_prezzo(listino_tot) or '0'}"
        )

    def _elimina_selezionato(self):
        """Elimina il tappeto selezionato."""
        selezione = self.sheet.get_currently_selected()
        if not selezione:
            messagebox.showwarning("Attenzione", "Seleziona un tappeto da eliminare.")
            return

        try:
            riga_idx = selezione.row if hasattr(selezione, "row") else selezione[0]
            riga_idx = int(riga_idx)
        except (TypeError, IndexError, ValueError):
            messagebox.showwarning("Attenzione", "Selezione non valida.")
            return

        if riga_idx < 0 or riga_idx >= len(self.df_visualizzato):
            messagebox.showwarning("Attenzione", "Selezione non valida.")
            return

        codice = str(self.df_visualizzato.iloc[riga_idx]["Nr"])
        nome = self.df_visualizzato.iloc[riga_idx]["Nome"]

        risposta = messagebox.askyesno(
            "Conferma Eliminazione",
            f"Sei sicuro di voler eliminare definitivamente il tappeto:\n\n"
            f"Codice: {codice}\nNome: {nome}"
        )

        if risposta:
            if self.gestore_db.elimina_tappeto(codice):
                messagebox.showinfo("Completato", "Tappeto eliminato con successo.")
                self.aggiorna_dati()

    def _esporta_stampa(self):
        """Esporta il dataframe filtrato per la stampa."""
        if self.df_visualizzato.empty:
            messagebox.showwarning("Attenzione", "Nessun dato da esportare.")
            return

        percorso = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("File Excel", "*.xlsx")],
            title="Salva elenco per stampa"
        )

        if percorso:
            try:
                # Calcola i totali sui valori NUMERICI originali
                # (prima di qualsiasi formattazione/rimozione di colonne)
                totali = {
                    "pezzi":   len(self.df_visualizzato),
                    "mq":      float(self.df_visualizzato["MQ"].sum()),
                    "costo":   float(self.df_visualizzato["Costo"].sum()),
                    "listino": float(self.df_visualizzato["Listino"].sum()),
                    "incasso": float(self.df_visualizzato["Prezzo Vendita Effettivo"].sum()),
                }

                # Esclude Data Vendita e Note dalla stampa
                colonne_da_escludere = ["Data Vendita", "Note"]
                df_stampa = self.df_visualizzato.drop(
                    columns=[c for c in colonne_da_escludere
                             if c in self.df_visualizzato.columns]
                ).reset_index(drop=True)

                # Aggiunge numero progressivo identificativo come prima colonna
                df_stampa.insert(0, "#", range(1, len(df_stampa) + 1))

                df_stampa.to_excel(percorso, index=False, engine="openpyxl")
                formatta_excel_per_stampa(percorso, totali=totali)
                messagebox.showinfo("Completato", f"File esportato con successo:\n{percorso}")
            except Exception as e:
                messagebox.showerror("Errore", f"Errore durante l'esportazione:\n{e}")

    def aggiorna_dati(self):
        """Aggiorna i dati dal database."""
        self.df_visualizzato = ordina_df_per_nr(self.gestore_db.df.copy())
        self._aggiorna_tabella()


class FrameAggiungiTappeto(ctk.CTkFrame):
    """Frame per la sezione Aggiungi Tappeto."""

    def __init__(self, master, gestore_db: GestoreDatabase, callback_aggiornamento):
        super().__init__(master, fg_color="transparent")
        self.gestore_db = gestore_db
        self.callback_aggiornamento = callback_aggiornamento
        self._crea_interfaccia()

    def _crea_interfaccia(self):
        """Crea tutti gli elementi dell'interfaccia aggiungi tappeto."""
        ctk.CTkLabel(
            self, text="Aggiungi Nuovo Tappeto",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(pady=(20, 10))

        self.scroll_frame = ctk.CTkScrollableFrame(self)
        self.scroll_frame.pack(fill="both", expand=True, padx=40, pady=10)

        self.inputs = {}

        campi = [
            ("Nr",          "entry",   True,  "Es. T001"),
            ("Nome",        "entry",   True,  "Nome del tappeto"),
            ("Provenienza", "combo",   False, None),
            ("Misura",      "entry",   True,  "Es. 200x150 (cm x cm)"),
            ("UM",          "combo",   False, None),
            ("Epoca",       "combo",   False, None),
            ("Qualita",     "combo",   False, None),
            ("Disegno",     "combo",   False, None),
            ("Colore",      "entry",   False, "Es. Rosso, Blu..."),
            ("Fornitore",   "entry",   False, "Nome fornitore"),
            ("Costo",       "entry",   False, "Es. 150"),
            ("Listino",     "entry",   False, "Es. 300"),
            ("Note",        "textbox", False, None),
        ]

        for i, (nome, tipo, obbligatorio, placeholder) in enumerate(campi):
            riga = i // 2
            colonna = i % 2

            frame_campo = ctk.CTkFrame(self.scroll_frame, fg_color="transparent")
            frame_campo.grid(row=riga, column=colonna, padx=20, pady=8, sticky="ew")

            label_text = f"{nome}*:" if obbligatorio else f"{nome}:"
            ctk.CTkLabel(
                frame_campo, text=label_text, width=100, anchor="w",
                font=ctk.CTkFont(weight="bold" if obbligatorio else "normal")
            ).pack(anchor="w")

            if tipo == "entry":
                kwargs = {"width": 250}
                if placeholder:
                    kwargs["placeholder_text"] = placeholder
                widget = ctk.CTkEntry(frame_campo, **kwargs)
                if nome == "Misura":
                    widget.bind("<KeyRelease>", self._aggiorna_anteprima_mq)
            elif tipo == "combo":
                if nome == "Epoca":
                    valori = OPZIONI_EPOCA
                elif nome == "Qualita":
                    valori = OPZIONI_QUALITA
                elif nome == "Provenienza":
                    valori = OPZIONI_PROVENIENZA
                elif nome == "UM":
                    valori = OPZIONI_UM
                else:
                    valori = OPZIONI_DISEGNO
                widget = ctk.CTkComboBox(frame_campo, values=valori, width=250)
                widget.set("")
            else:
                widget = ctk.CTkTextbox(frame_campo, width=250, height=80)

            widget.pack(anchor="w")
            self.inputs[nome] = widget

            if nome == "Misura":
                self.label_mq_preview = ctk.CTkLabel(
                    frame_campo,
                    text="MQ calcolati: —",
                    font=ctk.CTkFont(size=12),
                    text_color="#1f78b4"
                )
                self.label_mq_preview.pack(anchor="w", pady=(3, 0))

        self.scroll_frame.columnconfigure(0, weight=1)
        self.scroll_frame.columnconfigure(1, weight=1)

        frame_pulsanti = ctk.CTkFrame(self, fg_color="transparent")
        frame_pulsanti.pack(pady=15)

        ctk.CTkButton(
            frame_pulsanti, text="Salva Tappeto",
            command=self._salva_tappeto, width=200, height=40
        ).pack(side="left", padx=10)

        ctk.CTkButton(
            frame_pulsanti, text="Pulisci Form",
            fg_color="gray", hover_color="#555555",
            command=self._pulisci_form, width=140, height=40
        ).pack(side="left", padx=10)

    def _aggiorna_anteprima_mq(self, event=None):
        """Aggiorna il label anteprima MQ mentre si digita la misura."""
        misura = self.inputs["Misura"].get().strip()
        mq = calcola_mq_da_misura(misura)
        if mq > 0:
            self.label_mq_preview.configure(
                text=f"MQ calcolati: {mq:.2f} m²", text_color="#2d8a4e"
            )
        elif misura:
            self.label_mq_preview.configure(
                text="Formato non riconosciuto (usa: 200x150)",
                text_color="#dc3545"
            )
        else:
            self.label_mq_preview.configure(
                text="MQ calcolati: —", text_color="#1f78b4"
            )

    def _salva_tappeto(self):
        """Salva il nuovo tappeto nel database."""
        nr = self.inputs["Nr"].get().strip()
        nome = self.inputs["Nome"].get().strip()
        misura = self.inputs["Misura"].get().strip()

        if not nr or not nome or not misura:
            messagebox.showwarning(
                "Campi Obbligatori",
                "I campi Nr, Nome e Misura sono obbligatori."
            )
            return

        costo_str = self.inputs["Costo"].get().strip()
        listino_str = self.inputs["Listino"].get().strip()

        try:
            costo = float(costo_str.replace(",", ".")) if costo_str else 0
        except ValueError:
            messagebox.showwarning("Errore", "Il campo Costo deve contenere un valore numerico.")
            return

        try:
            listino = float(listino_str.replace(",", ".")) if listino_str else 0
        except ValueError:
            messagebox.showwarning("Errore", "Il campo Listino deve contenere un valore numerico.")
            return

        mq = calcola_mq_da_misura(misura)
        note = self.inputs["Note"].get("1.0", "end-1c").strip()

        dati = {
            "Stato": "Disponibile",
            "Nr": nr,
            "Nome": nome,
            "Provenienza": self.inputs["Provenienza"].get().strip(),
            "Misura": misura,
            "MQ": mq,
            "UM": self.inputs["UM"].get().strip(),
            "Epoca": self.inputs["Epoca"].get(),
            "Qualita": self.inputs["Qualita"].get(),
            "Disegno": self.inputs["Disegno"].get(),
            "Colore": self.inputs["Colore"].get().strip(),
            "Fornitore": self.inputs["Fornitore"].get().strip(),
            "Costo": costo,
            "Listino": listino,
            "Prezzo Vendita Effettivo": 0,
            "Data Vendita": "",
            "Note": note
        }

        if self.gestore_db.aggiungi_tappeto(dati):
            messagebox.showinfo("Completato", "Tappeto aggiunto con successo.")
            self._pulisci_form()
            self.callback_aggiornamento()

    def _pulisci_form(self):
        """Pulisce tutti i campi del form."""
        for nome, widget in self.inputs.items():
            if isinstance(widget, ctk.CTkEntry):
                widget.delete(0, "end")
            elif isinstance(widget, ctk.CTkComboBox):
                widget.set("")
            elif isinstance(widget, ctk.CTkTextbox):
                widget.delete("1.0", "end")
        self.label_mq_preview.configure(text="MQ calcolati: —", text_color="#1f78b4")


class FrameRegistraVendita(ctk.CTkFrame):
    """Frame per la sezione Registra Vendita."""

    def __init__(self, master, gestore_db: GestoreDatabase, callback_aggiornamento):
        super().__init__(master, fg_color="transparent")
        self.gestore_db = gestore_db
        self.callback_aggiornamento = callback_aggiornamento
        self._crea_interfaccia()

    def _crea_interfaccia(self):
        """Crea tutti gli elementi dell'interfaccia registra vendita."""
        ctk.CTkLabel(
            self, text="Registra Vendita",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(pady=(20, 20))

        frame_layout = ctk.CTkFrame(self, fg_color="transparent")
        frame_layout.pack(fill="both", expand=True, padx=40, pady=10)
        frame_layout.columnconfigure(0, weight=1)
        frame_layout.columnconfigure(1, weight=1)

        # Colonna sinistra — form vendita
        frame_form = ctk.CTkFrame(frame_layout)
        frame_form.grid(row=0, column=0, padx=(0, 10), pady=0, sticky="nsew")

        ctk.CTkLabel(
            frame_form, text="Seleziona Tappeto Disponibile:",
            font=ctk.CTkFont(size=14)
        ).pack(pady=(20, 5), padx=20)

        # Autocomplete al posto della ComboBox statica
        self.campo_ricerca = AutocompleteEntry(
            frame_form, width=380,
            placeholder="Scrivi Nr o nome...",
            on_select=self._mostra_dettagli_tappeto
        )
        self.campo_ricerca.pack(pady=10, padx=20, fill="x")

        ctk.CTkLabel(
            frame_form, text="Prezzo Vendita Effettivo (EUR):",
            font=ctk.CTkFont(size=14)
        ).pack(pady=(20, 5), padx=20)

        self.entry_prezzo = ctk.CTkEntry(
            frame_form, width=200, placeholder_text="Es. 350"
        )
        self.entry_prezzo.pack(pady=10)
        self.entry_prezzo.bind("<Return>", lambda e: self._conferma_vendita())

        ctk.CTkButton(
            frame_form, text="Conferma Vendita",
            command=self._conferma_vendita,
            width=200, height=40
        ).pack(pady=30)

        # Colonna destra — scheda dettagli tappeto selezionato
        self.frame_dettagli = ctk.CTkFrame(frame_layout)
        self.frame_dettagli.grid(row=0, column=1, padx=(10, 0), pady=0, sticky="nsew")

        ctk.CTkLabel(
            self.frame_dettagli, text="Scheda Tappeto Selezionato",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(pady=(15, 10), padx=20)

        self.labels_dettagli = {}
        campi_dettaglio = [
            "Nome", "Misura", "MQ", "Provenienza",
            "Fornitore", "Listino", "Qualita", "Note"
        ]
        for campo in campi_dettaglio:
            frame_riga = ctk.CTkFrame(self.frame_dettagli, fg_color="transparent")
            frame_riga.pack(fill="x", padx=20, pady=3)
            ctk.CTkLabel(
                frame_riga, text=f"{campo}:", width=80, anchor="w",
                font=ctk.CTkFont(size=12), text_color="gray"
            ).pack(side="left")
            lbl = ctk.CTkLabel(
                frame_riga, text="—", anchor="w", font=ctk.CTkFont(size=12)
            )
            lbl.pack(side="left", padx=5)
            self.labels_dettagli[campo] = lbl

        self.aggiorna_lista_tappeti()

    def _mostra_dettagli_tappeto(self, selezione: str = None):
        """Popola la scheda con i dati del tappeto selezionato."""
        if selezione is None:
            selezione = self.campo_ricerca.get()
        if not selezione:
            for lbl in self.labels_dettagli.values():
                lbl.configure(text="—", text_color=("gray10", "gray90"))
            return

        # Estrai il codice (formato "NR - Nome" o solo "NR")
        codice = selezione.split(" - ")[0].strip()
        disponibili = self.gestore_db.ottieni_disponibili()
        righe = disponibili[disponibili["Nr"].astype(str) == str(codice)]
        if righe.empty:
            return

        row = righe.iloc[0]
        self.labels_dettagli["Nome"].configure(text=str(row.get("Nome", "—") or "—"))
        self.labels_dettagli["Misura"].configure(text=str(row.get("Misura", "—") or "—"))

        mq = row.get("MQ", 0)
        self.labels_dettagli["MQ"].configure(
            text=f"{mq:.2f} m²" if mq else "—",
            text_color="#1f78b4" if mq else ("gray10", "gray90")
        )
        self.labels_dettagli["Provenienza"].configure(
            text=str(row.get("Provenienza", "—") or "—")
        )
        self.labels_dettagli["Fornitore"].configure(
            text=str(row.get("Fornitore", "—") or "—")
        )
        listino = row.get("Listino", 0)
        self.labels_dettagli["Listino"].configure(
            text=formatta_prezzo(listino) if listino else "—",
            text_color="#1f78b4" if listino else ("gray10", "gray90")
        )
        self.labels_dettagli["Qualita"].configure(
            text=str(row.get("Qualita", "—") or "—")
        )
        note = str(row.get("Note", "") or "")
        self.labels_dettagli["Note"].configure(
            text=(note[:50] + "...") if len(note) > 50 else (note or "—")
        )

    def aggiorna_lista_tappeti(self):
        """Aggiorna la lista dei tappeti disponibili nell'autocomplete."""
        disponibili = self.gestore_db.ottieni_disponibili()

        if disponibili.empty:
            self.campo_ricerca.set_values([])
            self.campo_ricerca.set("")
            for lbl in self.labels_dettagli.values():
                lbl.configure(text="—")
        else:
            opzioni = [
                f"{row['Nr']} - {row['Nome']}"
                for _, row in disponibili.iterrows()
            ]
            self.campo_ricerca.set_values(opzioni)
            # Non pre-seleziona alcun tappeto: il campo resta vuoto finché
            # l'utente non digita o sceglie esplicitamente.
            self.campo_ricerca.set("")
            for lbl in self.labels_dettagli.values():
                lbl.configure(text="—")

    def _conferma_vendita(self):
        """Conferma e registra la vendita."""
        selezione = self.campo_ricerca.get()

        if not selezione:
            messagebox.showwarning("Attenzione", "Seleziona un tappeto da vendere.")
            return

        # Estrai codice dal formato "NR - Nome"
        codice = selezione.split(" - ")[0].strip()

        # Verifica che il codice esista tra i disponibili
        disponibili = self.gestore_db.ottieni_disponibili()
        match_codice = disponibili[disponibili["Nr"].astype(str) == str(codice)]

        if match_codice.empty:
            # Prova a cercare per nome (se l'utente ha scritto solo il nome)
            match_nome = disponibili[
                disponibili["Nome"].astype(str).str.lower() == selezione.lower()
            ]
            if match_nome.empty:
                messagebox.showwarning(
                    "Attenzione",
                    "Tappeto non trovato tra i disponibili.\n"
                    "Seleziona un elemento dalla lista a tendina."
                )
                return
            codice = str(match_nome.iloc[0]["Nr"])

        prezzo_str = self.entry_prezzo.get().strip()
        if not prezzo_str:
            messagebox.showwarning("Attenzione", "Inserisci il prezzo di vendita effettivo.")
            return

        try:
            prezzo = float(prezzo_str.replace(",", "."))
        except ValueError:
            messagebox.showwarning("Errore", "Il prezzo deve essere un valore numerico.")
            return

        if prezzo < 0:
            messagebox.showwarning("Errore", "Il prezzo non puo essere negativo.")
            return

        risposta = messagebox.askyesno(
            "Conferma Vendita",
            f"Confermi la vendita del tappeto:\n\n"
            f"Codice: {codice}\n"
            f"Prezzo: {formatta_prezzo(prezzo) or str(prezzo)}"
        )

        if risposta:
            if self.gestore_db.registra_vendita(codice, prezzo):
                messagebox.showinfo("Completato", "Vendita registrata con successo.")
                self.entry_prezzo.delete(0, "end")
                self.aggiorna_lista_tappeti()
                self.callback_aggiornamento()


class AppMagazzino(ctk.CTk):
    """Classe principale dell'applicazione."""

    _BTN_ATTIVO  = ["#1A5C9A", "#164F8A"]
    _BTN_NORMALE = ["#3B8ED0", "#1F6AA5"]

    def __init__(self):
        super().__init__()

        self.title("Gestionale Magazzino Tappeti")
        self.geometry("1280x750")
        self.minsize(1000, 600)

        self.gestore_db = GestoreDatabase(FILE_DATABASE)
        self._sezione_attiva = "dashboard"
        self._crea_layout()

    def _crea_layout(self):
        """Crea il layout principale con sidebar e area contenuto."""
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=0)

        # Sidebar
        self.sidebar = ctk.CTkFrame(self, width=200, corner_radius=0)
        self.sidebar.grid(row=0, column=0, rowspan=2, sticky="nsew")
        self.sidebar.grid_rowconfigure(5, weight=1)

        ctk.CTkLabel(
            self.sidebar, text="Magazzino\nTappeti",
            font=ctk.CTkFont(size=20, weight="bold")
        ).grid(row=0, column=0, padx=20, pady=(20, 30))

        self._pulsanti_nav = {}
        nav_items = [
            ("dashboard", "Dashboard",        1),
            ("cerca",     "Cerca / Elenco",   2),
            ("aggiungi",  "Aggiungi Tappeto",  3),
            ("vendita",   "Registra Vendita",  4),
        ]
        for nome, testo, riga in nav_items:
            btn = ctk.CTkButton(
                self.sidebar, text=testo, anchor="w",
                command=lambda n=nome: self._mostra_sezione(n)
            )
            btn.grid(row=riga, column=0, padx=15, pady=5, sticky="ew")
            self._pulsanti_nav[nome] = btn

        # Area contenuto principale
        self.frame_contenuto = ctk.CTkFrame(self, fg_color="transparent")
        self.frame_contenuto.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        self.frame_contenuto.grid_columnconfigure(0, weight=1)
        self.frame_contenuto.grid_rowconfigure(0, weight=1)

        # Barra di stato
        self.frame_statusbar = ctk.CTkFrame(self, height=26, corner_radius=0)
        self.frame_statusbar.grid(row=1, column=0, columnspan=2, sticky="ew")

        self.label_status_db = ctk.CTkLabel(
            self.frame_statusbar,
            text=f"Database: {FILE_DATABASE}",
            font=ctk.CTkFont(size=11), text_color="gray", anchor="w"
        )
        self.label_status_db.pack(side="left", padx=12)

        self.label_status_records = ctk.CTkLabel(
            self.frame_statusbar, text="",
            font=ctk.CTkFont(size=11), text_color="gray"
        )
        self.label_status_records.pack(side="right", padx=12)

        # Crea le sezioni
        self.sezioni = {}
        self.sezioni["dashboard"] = FrameDashboard(self.frame_contenuto, self.gestore_db)
        self.sezioni["cerca"]     = FrameCercaElenco(self.frame_contenuto, self.gestore_db)
        self.sezioni["aggiungi"]  = FrameAggiungiTappeto(
            self.frame_contenuto, self.gestore_db, self._aggiorna_tutte_sezioni
        )
        self.sezioni["vendita"]   = FrameRegistraVendita(
            self.frame_contenuto, self.gestore_db, self._aggiorna_tutte_sezioni
        )

        for sezione in self.sezioni.values():
            sezione.grid(row=0, column=0, sticky="nsew")

        self._mostra_sezione("dashboard")

    def _aggiorna_status_bar(self):
        """Aggiorna i contatori nella barra di stato."""
        totale = len(self.gestore_db.df)
        disponibili = len(self.gestore_db.ottieni_disponibili())
        venduti = totale - disponibili
        self.label_status_records.configure(
            text=f"Totale: {totale}  |  Disponibili: {disponibili}  |  Venduti: {venduti}"
        )

    def _mostra_sezione(self, nome_sezione: str):
        """Mostra la sezione specificata e nasconde le altre."""
        self._sezione_attiva = nome_sezione

        for nome, btn in self._pulsanti_nav.items():
            btn.configure(
                fg_color=self._BTN_ATTIVO if nome == nome_sezione else self._BTN_NORMALE
            )

        for nome, sezione in self.sezioni.items():
            if nome == nome_sezione:
                sezione.tkraise()
                if hasattr(sezione, "aggiorna_dati"):
                    sezione.aggiorna_dati()
                if hasattr(sezione, "aggiorna_lista_tappeti"):
                    sezione.aggiorna_lista_tappeti()

        self._aggiorna_status_bar()

    def _aggiorna_tutte_sezioni(self):
        """Callback per aggiornare tutte le sezioni dopo modifiche."""
        for sezione in self.sezioni.values():
            if hasattr(sezione, "aggiorna_dati"):
                sezione.aggiorna_dati()
            if hasattr(sezione, "aggiorna_lista_tappeti"):
                sezione.aggiorna_lista_tappeti()
        self._aggiorna_status_bar()


def main():
    """Funzione principale di avvio."""
    app = AppMagazzino()
    app.mainloop()


if __name__ == "__main__":
    main()
